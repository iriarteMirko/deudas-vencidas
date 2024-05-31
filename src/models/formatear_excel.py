from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from tkinter import messagebox
import openpyxl


def formatear_excel(excel_file):
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        ws.title = "DETALLE"
        
        fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        font_header = Font(name="Calibri", size=11, color="000000", bold=True)
        font_cells = Font(name="Calibri", size=11)
        border = Border(left=Side(style="thin"), 
                        right=Side(style="thin"), 
                        top=Side(style="thin"), 
                        bottom=Side(style="thin"))
        alignment = Alignment(vertical="center")
        
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = alignment
                cell.font = font_cells
                if cell.row == 1:
                    cell.fill = fill
                    cell.font = font_header
                    cell.alignment = Alignment(horizontal="center")
        
        column_widths = [11, 30, 9, 24, 14, 12, 15, 21, 28.5]
        for i, column_width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i+1)].width = column_width
        
        wb.save(excel_file)
    except Exception as ex:
        messagebox.showerror("Error", "Algo salió mal. Por favor, intente nuevamente.\nDetalles: " + str(ex))