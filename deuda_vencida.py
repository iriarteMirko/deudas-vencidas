from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from tkinter import messagebox
from customtkinter import *
from resource_path import resource_path
from conexion import *
import pandas as pd
import openpyxl
import warnings
import threading
import time

warnings.filterwarnings("ignore")

class App_DV():
    def deshabilitar_botones(self):
        self.boton_ejecutar.configure(state="disabled")
        self.boton_base.configure(state="disabled")
        self.boton_dacx.configure(state="disabled")
        self.combobox_analistas.configure(state="disabled")
    
    def habilitar_botones(self):
        self.boton_ejecutar.configure(state="normal")
        self.boton_base.configure(state="normal")
        self.boton_dacx.configure(state="normal")
        self.combobox_analistas.configure(state="normal")
    
    def verificar_thread(self, thread):
        if thread.is_alive():
            self.app.after(1000, self.verificar_thread, thread)
        else:
            self.habilitar_botones()
    
    def iniciar_tarea(self, action):
        self.deshabilitar_botones()
        if action == 1:
            thread = threading.Thread(target=self.ejecutar)
        else:
            return
        thread.start()
        self.app.after(1000, self.verificar_thread, thread)
    
    def formatear_excel(self, excel_file):
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            ws.title = "DETALLE"
            
            fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            font_header = Font(name="Calibri", size=11, color="000000", bold=True)
            font_cells = Font(name="Calibri", size=11)
            border = Border(left=Side(style="thin"), 
                            right=Side(style="thin"), 
                            top=Side(style="thin"), 
                            bottom=Side(style="thin"))
            alignment = Alignment(vertical="center")
            
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = border
                    cell.alignment = alignment
                    cell.font = font_cells
                    if cell.row == 1:
                        cell.fill = fill
                        cell.font = font_header
                        cell.alignment = Alignment(horizontal="center")
            
            column_widths = [10.5, 30, 8.5, 23, 13.5, 12, 14, 20, 27]
            for i, column_width in enumerate(column_widths):
                ws.column_dimensions[get_column_letter(i+1)].width = column_width
            
            wb.save(excel_file)
        except Exception as ex:
            messagebox.showerror("Error", "Algo salió mal. Por favor, intente nuevamente.\nDetalles: " + str(ex))
    
    def obtener_deudas_vencidas(self, base_path, dacxanalista_path, resultado_path):
        start = time.time()
        lista_estados = []
        variables = [
            (self.var_ope_con_mov, "OPERATIVO CON MOVIMIENTO"),
            (self.var_ope_sin_mov, "OPERATIVO SIN MOVIMIENTO"),
            (self.var_proc_liquidacion, "PROCESO DE LIQUIDACIÓN"),
            (self.var_proc_pre_resolucion, "PROCESO DE PRE RESOLUCION"),
            (self.var_proc_resolucion, "PROCESO DE RESOLUCIÓN"),
            (self.var_liquidado, "LIQUIDADO")
        ]
        for var, estado in variables:
            if var.get():
                lista_estados.append(estado)
        
        if len(lista_estados) == 0:
            messagebox.showerror("Error", "Por favor, seleccione al menos un estado.")
            return
        try:
            df_base = pd.read_excel(base_path)
            df_base.dropna(subset=["ACC","Cuenta"], inplace=True)
            df_base = df_base.reset_index(drop=True)
            df_base = df_base.rename(columns={"Importe en ML": "Importe"})
            columnas_deseadas = ["ACC", "Cuenta", "Demora", "Importe"]
            df_base = df_base[columnas_deseadas]
            df_base["Demora"] = df_base["Demora"].astype("Int64")
            df_base["Importe"] = df_base["Importe"].astype(float)
            df_base = df_base.reset_index(drop=True)
            
            df_dacxanalista = pd.read_excel(dacxanalista_path, sheet_name="Base_NUEVA")
            columnas_dacx = ["DEUDOR", "NOMBRE", "ANALISTA_ACT", "ESTADO"]
            df_dacxanalista = df_dacxanalista[columnas_dacx]
            
            if self.combobox_analistas.get() != "TODOS":
                df_dacxanalista = df_dacxanalista[df_dacxanalista["ANALISTA_ACT"] == self.combobox_analistas.get()]
            else:
                df_dacxanalista = df_dacxanalista[df_dacxanalista["ANALISTA_ACT"] != "SIN INFORMACION"]
            
            if len(lista_estados) > 0:
                df_dacxanalista = df_dacxanalista[df_dacxanalista["ESTADO"].isin(lista_estados)]
            
            lista_cartera = df_dacxanalista["DEUDOR"].tolist()
            df_base = df_base[df_base["Cuenta"].isin(lista_cartera)]
            
            df_base["Status"] = df_base["Importe"].apply(lambda x: "DEUDA" if x > 0 else "SALDOS A FAVOR")
            df_base["Tipo Deuda"] = df_base["Demora"].apply(lambda x: "CORRIENTE" if x <= 0 else "VENCIDA")
            df_base["Saldo Final"] = df_base.apply(lambda row: row["Importe"] 
                                                    if (row["Status"]=="DEUDA" and row["Tipo Deuda"]=="VENCIDA") 
                                                    else (row["Importe"] if row["Status"]=="SALDOS A FAVOR" else "NO"), 
                                                    axis=1)
            df_base = df_base[df_base["Saldo Final"] != "NO"]
            df_base = df_base.sort_values(by=["Cuenta"], ascending=[True])
            df_base = df_base.sort_values(by=["ACC"], ascending=[True])
            df_base = df_base.sort_values(by=["Demora"], ascending=[False])
            df_base = df_base.sort_values(by=["Cuenta"], ascending=[True])
            df_base = df_base.reset_index(drop=True)
            
            cuentas_verificadas = []
            ultima_fila = df_base.shape[0]
            for i in range(ultima_fila):
                cuenta_actual = df_base.loc[i, "Cuenta"]
                if cuenta_actual not in cuentas_verificadas:
                    cuentas_verificadas.append(cuenta_actual)
                    inicio = i
                if df_base.loc[i, "Status"] == "DEUDA":
                    saldoDeuda = df_base.loc[i, "Saldo Final"]
                    rango = df_base[df_base['Cuenta'] == cuenta_actual].shape[0]
                    for j in range(inicio, inicio+rango):
                        if (
                            df_base.loc[j, "Cuenta"] == cuenta_actual and 
                            df_base.loc[j, "ACC"] == df_base.loc[i, "ACC"] and 
                            df_base.loc[j, "Status"] == "SALDOS A FAVOR"
                            ):
                            saldoFavor = df_base.loc[j, "Saldo Final"]
                            montoCompensar = min(saldoDeuda, abs(saldoFavor))
                            df_base.loc[i, "Saldo Final"] = saldoDeuda - montoCompensar
                            df_base.loc[j, "Saldo Final"] = saldoFavor + montoCompensar
                            saldoDeuda = df_base.loc[i, "Saldo Final"]
            
            df_base = df_base[(df_base["Tipo Deuda"] == "VENCIDA") & (df_base["Status"] == "DEUDA")]
            df_base = df_base.reset_index(drop=True)
            grouped_df = df_base.groupby(["Cuenta", "ACC"]).agg({"Demora": "max", "Saldo Final": "sum"})
            
            df_final = grouped_df.reset_index()[["Cuenta", "ACC", "Saldo Final", "Demora"]]
            df_final = df_final.rename(columns={"Cuenta": "Cod Cliente", "ACC": "Área Ctrl", 
                                                "Saldo Final": "Deuda Vencida", "Demora": "Días Morosidad"})
            df_final = df_final.merge(df_dacxanalista[["DEUDOR", "NOMBRE", "ANALISTA_ACT", "ESTADO"]], 
                                        left_on="Cod Cliente", right_on="DEUDOR", how="left")
            df_final = df_final.rename(columns={"NOMBRE": "Razón Social", "ANALISTA_ACT": "Analista", "ESTADO": "Estado"})
            df_final = df_final.drop(columns=["DEUDOR"])
            areas_de_control = {
                "PE01": "Post-Pago",
                "PE02": "Pre-Pago",
                "PE03": "Tiempo Aire",
                "PE04": "Reintegro",
                "PE05": "Reestructura",
                "PE07": "Contado / Administrativas",
                "PE09": "Cargos Admtivos / Otros",
                "PE10": "Sim Card",
                "PE11": "Recarga Prepago",
                "PE12": "Recarga Física",
                "PE13": "Arrendamiento",
                "PE14": "Tel.Fija Inalamb.",
                "PE15": "Prendas",
                "PE16": "DTH",
                "PE17": "HFC"
            }
            df_final = df_final[df_final["Área Ctrl"].isin(areas_de_control.keys())]
            df_final["Producto"] = df_final["Área Ctrl"].apply(lambda x: areas_de_control.get(x))
            df_final["Código Pago"] = "33" + df_final["Área Ctrl"].str[-2:] + df_final["Cod Cliente"].astype(str)
            df_final = df_final[["Cod Cliente", "Razón Social", "Área Ctrl", "Producto", "Deuda Vencida", 
                                "Código Pago", "Días Morosidad", "Analista", "Estado"]]
            df_final["Deuda Vencida"] = df_final["Deuda Vencida"].astype(float).round(2)
            df_final = df_final[df_final["Deuda Vencida"] != 0]
            df_final = df_final.sort_values(by=["Área Ctrl"], ascending=[True])
            df_final = df_final.sort_values(by=["Cod Cliente"], ascending=[True])
            df_final = df_final.sort_values(by=["Deuda Vencida"], ascending=[False])
            df_final = df_final.sort_values(by=["Días Morosidad"], ascending=[False])
            df_final.to_excel(resultado_path, index=False)
            
            self.formatear_excel(resultado_path)
            end = time.time()
            messagebox.showinfo("Éxito", "Registros encontrados: " + str(df_final.shape[0]) + 
                                "\nTiempo de ejecución: " + str(round(end-start,2)) + " segundos.")
            os.startfile(resultado_path)
        except Exception as ex:
            messagebox.showerror("Error", "Algo salió mal. Por favor, intente nuevamente.\nDetalles: " + str(ex))
    
    def seleccionar_base(self):
        archivo_excel = filedialog.askopenfilename(
            initialdir="/",
            title="Seleccionar archivo BASE",
            filetypes=(("Archivos de Excel", "*.xlsx"), ("Todos los archivos", "*.*"))
        )
        directorio_base = os.path.dirname(archivo_excel)
        base_path = archivo_excel
        resultado_path = directorio_base+"/DEUDAS_VENCIDAS.xlsx"
        
        query1 = ("""UPDATE RUTAS
                    SET BASE == '""" + base_path + """'
                    WHERE ID == 0""")
        query2 = ("""UPDATE RUTAS
                    SET RESULTADO == '""" + resultado_path + """'
                    WHERE ID == 0""")
        conexion = conexionSQLite()
        try:
            cursor = conexion.cursor()
            cursor.execute(query1)
            conexion.commit()
            cursor.execute(query2)
            conexion.commit()
        except Exception as ex:
            messagebox.showerror("Error", str(ex))
        finally:
            cursor.close()
            conexion.close
    
    def seleccionar_dacxanalista(self):
        archivo_excel = filedialog.askopenfilename(
            initialdir="/",
            title="Seleccionar archivo DACxANALISTA",
            filetypes=(("Archivos de Excel", "*.xlsx"), ("Todos los archivos", "*.*"))
        )
        dacxanalista_path = archivo_excel
        
        query = ("""UPDATE RUTAS
                    SET DACXANALISTA == '""" + dacxanalista_path + """'
                    WHERE ID == 0""")
        conexion = conexionSQLite()
        try:
            cursor = conexion.cursor()
            cursor.execute(query)
            conexion.commit()
        except Exception as ex:
            messagebox.showerror("Error", str(ex))
        finally:
            cursor.close()
            conexion.close
    
    def ejecutar(self):
        self.progressbar.start()
        query = """SELECT * FROM RUTAS WHERE ID == 0"""
        try:
            datos = ejecutar_query(query)
            ruta_base = datos[0][1]
            ruta_dacxa = datos[0][2]
            ruta_resultado = datos[0][3]
            if ruta_base is None or ruta_dacxa is None or ruta_resultado is None:
                messagebox.showerror("Error", "Por favor, configure las rutas de los archivos.")
            elif not os.path.exists(ruta_base):
                messagebox.showerror("Error", "No se encontraró el archivo BASE en la ruta especificada.")
            elif not os.path.exists(ruta_dacxa):
                messagebox.showerror("Error", "No se encontraró el archivo DACxANALISTA en la ruta especificada.")
            else:
                self.obtener_deudas_vencidas(ruta_base, ruta_dacxa, ruta_resultado)
        except Exception as ex:
            messagebox.showerror("Error", str(ex))
        finally:
            self.progressbar.stop()
    
    def crear_app(self):        
        self.app = CTk()
        self.app.title("Deudas Vencidas")
        self.app.iconbitmap(resource_path("./images/icono.ico"))
        self.app.resizable(False, False)
        set_appearance_mode("light")
        
        main_frame = CTkFrame(self.app)
        main_frame.pack_propagate(0)
        main_frame.pack(fill="both", expand=True)
        
        frame_title = CTkFrame(main_frame)
        frame_title.grid(row=0, column=0, columnspan=2, padx=(20, 20), pady=(20, 0), sticky="nsew")
        
        titulo = CTkLabel(frame_title, text="  DEUDAS VENCIDAS  ", font=("Arial",25,"bold"))
        titulo.pack(fill="both", expand=True, ipady=20, anchor="center")
        
        frame_base = CTkFrame(main_frame)
        frame_base.grid(row=1, column=0, padx=(20, 10), pady=(20, 0), sticky="nsew")
        
        ruta_base = CTkLabel(frame_base, text="Ruta BASE", font=("Calibri",17,"bold"))
        ruta_base.pack(padx=(20, 20), pady=(5, 0), fill="both", expand=True, anchor="center", side="top")
        self.boton_base = CTkButton(frame_base, text="Seleccionar", font=("Calibri",17), text_color="black",
                                fg_color="transparent", border_color="#d11515", border_width=3, hover_color="#d11515", 
                                width=25, corner_radius=25, command=lambda: self.seleccionar_base())
        self.boton_base.pack(padx=(20, 20), pady=(0, 15), fill="both", anchor="center", side="bottom")
        
        frame_dacx = CTkFrame(main_frame)
        frame_dacx.grid(row=1, column=1, padx=(10, 20), pady=(20, 0), sticky="nsew")
        
        ruta_dacxa = CTkLabel(frame_dacx, text="Ruta DACxAnalista", font=("Calibri",17,"bold"))
        ruta_dacxa.pack(padx=(20, 20), pady=(5, 0), fill="both", expand=True, anchor="center", side="top")
        self.boton_dacx = CTkButton(frame_dacx, text="Seleccionar", font=("Calibri",17), text_color="black",
                                fg_color="transparent", border_color="#d11515", border_width=3, hover_color="#d11515", 
                                width=25, corner_radius=25, command=lambda: self.seleccionar_dacxanalista())
        self.boton_dacx.pack(padx=(20, 20), pady=(0, 15), fill="both", anchor="center", side="bottom")
        
        frame_analista = CTkFrame(main_frame)
        frame_analista.grid(row=2, column=0, columnspan=2, padx=(20, 20), pady=(20, 0), sticky="nsew")
        
        analistas = ["TODOS", "RAQUEL CAYETANO", "YOLANDA OLIVA", "DIEGO RODRIGUEZ", "JOSE LUIS VALVERDE", 
                    "JUAN CARLOS HUATAY", "WALTER LOPEZ", "REGION NORTE", "REGION SUR"]
        label_analista = CTkLabel(frame_analista, text="Analista Actual: ", font=("Calibri",18,"bold"))
        label_analista.pack(padx=(20, 0), pady=(15, 15), fill="both", expand=True, anchor="w", side="left")
        self.combobox_analistas = CTkComboBox(frame_analista, font=("Calibri",17), width=200, values=analistas, 
                                            state="readonly", border_color="#d11515")
        self.combobox_analistas.pack(padx=(0, 40), pady=(15, 15), fill="both", expand=True, anchor="w", side="right")
        self.combobox_analistas.set("TODOS")
        
        frame_estado = CTkFrame(main_frame)
        frame_estado.grid(row=3, column=0, columnspan=2, padx=(20, 20), pady=(20, 0), sticky="nsew")
        
        label_estado_dac = CTkLabel(frame_estado, text="Seleccionar Estados: ", font=("Calibri",18,"bold"))
        label_estado_dac.grid(row=0, column=0, columnspan=2, padx=(20, 20), pady=(10, 0), sticky="nsew")
        
        self.var_ope_con_mov = BooleanVar()
        self.var_ope_con_mov.set(True)
        ope_con_mov = CTkCheckBox(frame_estado, text="OP. CON MOVIMIENTO", font=("Calibri",17), 
                                    border_color="#d11515", border_width=2, fg_color="#d11515", 
                                    hover_color="#d11515", variable=self.var_ope_con_mov)
        ope_con_mov.grid(row=1, column=0, padx=(20, 10), pady=(10, 0), sticky="nsew")
        
        self.var_ope_sin_mov = BooleanVar()
        self.var_ope_sin_mov.set(True)
        ope_sin_mov = CTkCheckBox(frame_estado, text="OP. SIN MOVIMIENTO", font=("Calibri",17), 
                                    border_color="#d11515", border_width=2, fg_color="#d11515", 
                                    hover_color="#d11515", variable=self.var_ope_sin_mov)
        ope_sin_mov.grid(row=1, column=1, padx=(10, 20), pady=(10, 0), sticky="nsew")
        
        self.var_proc_liquidacion = BooleanVar()
        self.var_proc_liquidacion.set(False)
        proc_liquidacion = CTkCheckBox(frame_estado, text="PROC. LIQUIDACION", font=("Calibri",17), 
                                        border_color="#d11515", border_width=2, fg_color="#d11515", 
                                        hover_color="#d11515", variable=self.var_proc_liquidacion)
        proc_liquidacion.grid(row=2, column=0, padx=(20, 10), pady=(10, 0), sticky="nsew")
        
        self.var_proc_resolucion = BooleanVar()
        self.var_proc_resolucion.set(False)
        proc_resolucion = CTkCheckBox(frame_estado, text="PROC. RESOLUCION", font=("Calibri",17), 
                                            border_color="#d11515", border_width=2, fg_color="#d11515", 
                                            hover_color="#d11515", variable=self.var_proc_resolucion)
        proc_resolucion.grid(row=2, column=1, padx=(10, 20), pady=(10, 0), sticky="nsew")
        
        self.var_proc_pre_resolucion = BooleanVar()
        self.var_proc_pre_resolucion.set(False)
        proc_pre_resolucion = CTkCheckBox(frame_estado, text="PROC. PRE RESOLUCION", font=("Calibri",17), 
                                        border_color="#d11515", border_width=2, fg_color="#d11515", 
                                        hover_color="#d11515", variable=self.var_proc_pre_resolucion)
        proc_pre_resolucion.grid(row=3, column=0, padx=(20, 10), pady=(10, 20), sticky="nsew")
        
        self.var_liquidado = BooleanVar()
        self.var_liquidado.set(False)
        liquidado = CTkCheckBox(frame_estado, text="LIQUIDADO", font=("Calibri",17), border_color="#d11515", 
                                border_width=2, fg_color="#d11515", hover_color="#d11515", 
                                variable=self.var_liquidado)
        liquidado.grid(row=3, column=1, padx=(10, 20), pady=(10, 20), sticky="nsew")
        
        self.boton_ejecutar = CTkButton(main_frame, text="EJECUTAR", text_color="black", font=("Calibri",25,"bold"), 
                                    border_color="black", border_width=3, fg_color="gray", 
                                    hover_color="red", command=lambda: self.iniciar_tarea(1))
        self.boton_ejecutar.grid(row=4, column=0, columnspan=2, ipady=20, padx=(20, 20), pady=(20, 0), sticky="nsew")
        
        self.progressbar = CTkProgressBar(main_frame, mode="indeterminate", orientation="horizontal", 
                                        progress_color="#d11515", height=10, border_width=0)
        self.progressbar.grid(row=5, column=0, columnspan=2, padx=(20, 20), pady=(20, 20), sticky="nsew")
        
        self.app.mainloop()

def main():
    app = App_DV()
    app.crear_app()

if __name__ == "__main__":
    main()